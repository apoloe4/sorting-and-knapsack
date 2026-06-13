"""
0-1背包实验图表生成 + Excel数据文件
"""
import csv
import os
import math
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import numpy as np

plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei', 'SimSun']
plt.rcParams['axes.unicode_minus'] = False

BASE = r"C:\Users\f1969\WorkBuddy\2026-06-02-19-32-36"

# ---- 读取数据 ----
csv_path = os.path.join(BASE, 'knapsack_results.csv')
data = {}  # data[algo][cap] = {n: time_ms}
algos_in_file = set()
caps_in_file  = set()

with open(csv_path, newline='', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        algo = row['algo']
        n    = int(row['n'])
        cap  = int(row['capacity'])
        try:
            t = float(row['time_ms'])
        except:
            t = -1.0
        if algo not in data:   data[algo] = {}
        if cap  not in data[algo]: data[algo][cap] = {}
        data[algo][cap][n] = t if t > 0 else None
        algos_in_file.add(algo)
        caps_in_file.add(cap)

SIZES = [1000, 2000, 3000, 4000, 5000, 6000, 7000, 8000, 9000, 10000,
         20000, 40000, 80000, 160000, 320000]
CAPACITIES = sorted(caps_in_file)
ALGOS = ['dp', 'greedy', 'backtrack']

ALGO_LABELS = {'dp': '动态规划法 DP', 'greedy': '贪心法 Greedy', 'backtrack': '回溯法 Backtrack'}
ALGO_COLORS = {'dp': '#E74C3C', 'greedy': '#2ECC71', 'backtrack': '#3498DB'}
ALGO_MARKERS = {'dp': 'o', 'greedy': 's', 'backtrack': '^'}

# ---- 图1: 三种算法在三种容量下的执行时间对比（3子图）----
fig1, axes = plt.subplots(1, 3, figsize=(18, 6), sharey=False)
fig1.suptitle('0-1背包问题：三种算法执行时间对比', fontsize=15, fontweight='bold')

for idx, cap in enumerate(CAPACITIES):
    ax = axes[idx]
    for algo in ALGOS:
        xs, ys = [], []
        for n in SIZES:
            t = data.get(algo, {}).get(cap, {}).get(n)
            if t is not None:
                xs.append(n)
                ys.append(t)
        if xs:
            ax.plot(xs, ys, marker=ALGO_MARKERS[algo], color=ALGO_COLORS[algo],
                    linewidth=2, markersize=6, label=ALGO_LABELS[algo])

    ax.set_title(f'背包容量 C = {cap:,}', fontsize=12, fontweight='bold')
    ax.set_xlabel('物品数量 n', fontsize=11)
    ax.set_ylabel('执行时间 (ms)', fontsize=11)
    ax.legend(fontsize=9)
    ax.grid(True, linestyle='--', alpha=0.5)
    ax.xaxis.set_major_formatter(ticker.FuncFormatter(
        lambda x, _: f'{int(x/1000)}K' if x >= 1000 else str(int(x))))
    ax.yaxis.set_major_formatter(ticker.FuncFormatter(
        lambda x, _: f'{x/1000:.0f}s' if x >= 10000 else f'{x:.0f}ms'))

fig1.tight_layout()
fig1.savefig(os.path.join(BASE, 'knapsack_time_compare.png'), dpi=150, bbox_inches='tight')
print('图1 已保存: knapsack_time_compare.png')
plt.close(fig1)

# ---- 图2: 对数坐标，更清晰看贪心 vs 回溯 vs DP ----
fig2, axes2 = plt.subplots(1, 3, figsize=(18, 6), sharey=False)
fig2.suptitle('0-1背包问题：执行时间对比（对数坐标）', fontsize=15, fontweight='bold')

for idx, cap in enumerate(CAPACITIES):
    ax = axes2[idx]
    for algo in ALGOS:
        xs, ys = [], []
        for n in SIZES:
            t = data.get(algo, {}).get(cap, {}).get(n)
            if t is not None and t > 0:
                xs.append(n)
                ys.append(t)
        if xs:
            ax.plot(xs, ys, marker=ALGO_MARKERS[algo], color=ALGO_COLORS[algo],
                    linewidth=2, markersize=6, label=ALGO_LABELS[algo])

    ax.set_xscale('log')
    ax.set_yscale('log')
    ax.set_title(f'背包容量 C = {cap:,}', fontsize=12, fontweight='bold')
    ax.set_xlabel('物品数量 n', fontsize=11)
    ax.set_ylabel('执行时间 (ms)', fontsize=11)
    ax.legend(fontsize=9)
    ax.grid(True, which='both', linestyle='--', alpha=0.4)

fig2.tight_layout()
fig2.savefig(os.path.join(BASE, 'knapsack_time_loglog.png'), dpi=150, bbox_inches='tight')
print('图2 已保存: knapsack_time_loglog.png')
plt.close(fig2)

# ---- 图3: 蛮力法小规模时间增长（指数爆炸）----
brute_csv = os.path.join(BASE, 'knapsack_brute.csv')
brute_data = {}  # cap -> {n: time}
with open(brute_csv, newline='', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        n = int(row['n']); cap = int(row['capacity'])
        t = float(row['time_ms'])
        if cap not in brute_data: brute_data[cap] = {}
        brute_data[cap][n] = t

fig3, ax3 = plt.subplots(figsize=(9, 6))
brute_ns = [5, 10, 15, 20, 25]
for cap, c in zip([100, 500, 1000], ['#E74C3C', '#3498DB', '#2ECC71']):
    ys = [brute_data[cap].get(n, None) for n in brute_ns]
    valid = [(n, y) for n, y in zip(brute_ns, ys) if y is not None]
    if valid:
        xs2, ys2 = zip(*valid)
        ax3.plot(xs2, ys2, 'o-', color=c, linewidth=2, markersize=7, label=f'C={cap}')

# 叠加理论 2^n 曲线
ns_theory = np.array(brute_ns, dtype=float)
theory_base = brute_data[100].get(10, 0.027)
theory_coeff = theory_base / (2**10)
theory_vals = theory_coeff * 2**ns_theory
ax3.plot(ns_theory, theory_vals, '--', color='gray', linewidth=1.5, label='理论 O(2^n)')

ax3.set_yscale('log')
ax3.set_xlabel('物品数量 n', fontsize=12)
ax3.set_ylabel('执行时间 (ms，对数轴)', fontsize=12)
ax3.set_title('蛮力法执行时间（指数增长，n<=25）', fontsize=13, fontweight='bold')
ax3.legend(fontsize=10)
ax3.grid(True, which='both', linestyle='--', alpha=0.5)
ax3.set_xticks(brute_ns)
fig3.tight_layout()
fig3.savefig(os.path.join(BASE, 'knapsack_brute_time.png'), dpi=150, bbox_inches='tight')
print('图3 已保存: knapsack_brute_time.png')
plt.close(fig3)

# ---- 图4: DP法复杂度分析（固定n，改变C；固定C，改变n）----
fig4, axes4 = plt.subplots(1, 2, figsize=(14, 6))

# 左：固定C，改变n（DP应该线性于n）
ax4a = axes4[0]
for cap, c in zip(CAPACITIES, ['#E74C3C','#3498DB','#2ECC71']):
    xs, ys = [], []
    for n in SIZES:
        t = data.get('dp', {}).get(cap, {}).get(n)
        if t is not None:
            xs.append(n); ys.append(t)
    if xs:
        ax4a.plot(xs, ys, 'o-', color=c, linewidth=2, markersize=5, label=f'C={cap:,}')
ax4a.set_xlabel('物品数量 n', fontsize=12)
ax4a.set_ylabel('执行时间 (ms)', fontsize=12)
ax4a.set_title('DP法：不同容量下执行时间随n的变化', fontsize=12, fontweight='bold')
ax4a.legend(fontsize=10)
ax4a.grid(True, linestyle='--', alpha=0.5)
ax4a.xaxis.set_major_formatter(ticker.FuncFormatter(
    lambda x, _: f'{int(x/1000)}K' if x >= 1000 else str(int(x))))

# 右：固定n=10000，三种算法在三种容量下的时间
ax4b = axes4[1]
n_fixed = 10000
cap_labels = ['C=10,000', 'C=100,000', 'C=1,000,000']
x_pos = np.arange(3)
bar_width = 0.25
for i, algo in enumerate(ALGOS):
    heights = []
    for cap in CAPACITIES:
        t = data.get(algo, {}).get(cap, {}).get(n_fixed)
        heights.append(t if t else 0)
    ax4b.bar(x_pos + i * bar_width, heights, bar_width,
             label=ALGO_LABELS[algo], color=ALGO_COLORS[algo], alpha=0.85)

ax4b.set_xticks(x_pos + bar_width)
ax4b.set_xticklabels(cap_labels, fontsize=10)
ax4b.set_ylabel('执行时间 (ms)', fontsize=12)
ax4b.set_title(f'n={n_fixed:,} 时，三种算法在不同容量下的时间对比', fontsize=12, fontweight='bold')
ax4b.legend(fontsize=10)
ax4b.grid(True, axis='y', linestyle='--', alpha=0.5)
ax4b.set_yscale('log')

fig4.suptitle('DP法复杂度分析', fontsize=14, fontweight='bold')
fig4.tight_layout()
fig4.savefig(os.path.join(BASE, 'knapsack_dp_analysis.png'), dpi=150, bbox_inches='tight')
print('图4 已保存: knapsack_dp_analysis.png')
plt.close(fig4)

# ---- Excel 数据文件 ----
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    print('\n生成 Excel 文件...')

    wb = openpyxl.Workbook()

    # 工作表1：1000个物品的统计信息
    ws1 = wb.active
    ws1.title = '1000物品统计'

    # 生成物品数据（与C程序相同的随机种子42，n=1000）
    import random
    random.seed(42)
    items_1000 = []
    for i in range(1, 1001):
        w = random.randint(1, 100)
        cents = random.randint(10000, 100000)
        v = cents / 100.0
        items_1000.append((i, w, v))

    # 标题行
    header_fill = PatternFill('solid', fgColor='2980B9')
    header_font = Font(bold=True, color='FFFFFF')
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws1['A1'] = '物品编号'
    ws1['B1'] = '物品重量'
    ws1['C1'] = '物品价值'
    for col in ['A','B','C']:
        cell = ws1[f'{col}1']
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    for i, (item_id, w, v) in enumerate(items_1000):
        row = i + 2
        ws1.cell(row=row, column=1, value=item_id)
        ws1.cell(row=row, column=2, value=w)
        ws1.cell(row=row, column=3, value=round(v, 2))
        for col in range(1, 4):
            ws1.cell(row=row, column=col).border = border
            ws1.cell(row=row, column=col).alignment = Alignment(horizontal='center')

    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 14

    # 工作表2：实验结果汇总
    ws2 = wb.create_sheet('实验结果汇总')
    headers2 = ['算法', '物品数n', '背包容量C', '执行时间(ms)', '总价值', '总重量', '状态']
    for j, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=j, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    row2 = 2
    algo_labels_en = {'dp': '动态规划法', 'greedy': '贪心法', 'backtrack': '回溯法'}
    with open(csv_path, newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for rec in reader:
            algo = rec['algo']
            vals = [algo_labels_en.get(algo, algo),
                    int(rec['n']),
                    int(rec['capacity']),
                    float(rec['time_ms']) if float(rec['time_ms']) > 0 else '超时',
                    float(rec['total_value']) if float(rec['total_value']) > 0 else '-',
                    int(rec['total_weight']) if int(rec['total_weight']) > 0 else '-',
                    rec['status']]
            for j, val in enumerate(vals, 1):
                cell = ws2.cell(row=row2, column=j, value=val)
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            row2 += 1

    for col in range(1, 8):
        ws2.column_dimensions[get_column_letter(col)].width = 16

    # 工作表3：蛮力法小规模数据
    ws3 = wb.create_sheet('蛮力法小规模')
    headers3 = ['物品数n', '背包容量C', '执行时间(ms)', '最优总价值', '状态']
    for j, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=j, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    row3 = 2
    with open(brute_csv, newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for rec in reader:
            vals = [int(rec['n']), int(rec['capacity']),
                    float(rec['time_ms']), float(rec['value']), rec['status']]
            for j, val in enumerate(vals, 1):
                cell = ws3.cell(row=row3, column=j, value=val)
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            row3 += 1

    for col in range(1, 6):
        ws3.column_dimensions[get_column_letter(col)].width = 16

    xlsx_path = os.path.join(BASE, 'knapsack_data.xlsx')
    wb.save(xlsx_path)
    print(f'Excel 已保存: {xlsx_path}')

except ImportError:
    print('openpyxl 未安装，跳过Excel生成')
    xlsx_path = None

print('\n所有图表和数据文件生成完毕！')
